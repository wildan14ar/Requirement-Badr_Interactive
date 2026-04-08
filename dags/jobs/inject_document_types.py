"""
Script untuk inject/upsert data ke tabel official_document_types
Mode: UPSERT (insert jika belum ada, update jika sudah ada berdasarkan nama)
Data Source: Excel file (data/Master Data Jenis Naskah Dinas (official_document_type).xlsx)
"""

import psycopg2
from datetime import datetime
from ulid import ULID
from openpyxl import load_workbook
import json

# Database connection
DATABASE_URL = "postgresql://adhidms:adhidms2O2S@192.168.11.41:5432/postgres"

# Excel file path
EXCEL_FILE = r"data/Master Data Jenis Naskah Dinas (official_document_type).xlsx"

# Base URL for template files
TEMPLATE_BASE_URL = "https://dms-storage.adhi.co.id/template-naskah-dinas"


def generate_ulid():
    """Generate ULID baru"""
    return str(ULID())


def connect_db():
    """Koneksi ke database"""
    try:
        conn = psycopg2.connect(DATABASE_URL)
        return conn
    except Exception as e:
        print(f"❌ Error koneksi database: {e}")
        return None


def read_excel_data():
    """Baca data dari Excel file"""
    print(f"📂 Reading Excel file: {EXCEL_FILE}")
    
    workbook = load_workbook(EXCEL_FILE)
    sheet = workbook["Sheet1"]  # Main data sheet
    
    # Get all data
    data = list(sheet.values)
    
    if not data:
        print("❌ Excel file kosong!")
        return [], {}
    
    # Get column names (first row)
    columns = data[0]
    
    # Convert to list of dictionaries
    document_types = []
    for row_data in data[1:]:  # Skip header
        if row_data[0]:  # Skip if name is empty
            row_dict = dict(zip(columns, row_data))
            document_types.append(row_dict)
    
    # Read Sheet2 for formats
    formats_data = {}
    if "Formats" in workbook.sheetnames:
        sheet2 = workbook["Formats"]
        data2 = list(sheet2.values)
        
        if len(data2) >= 2:
            headers = data2[0]
            values = data2[1]
            
            for header, value in zip(headers, values):
                if value:
                    try:
                        parsed = json.loads(value)
                        formats_data[header] = parsed
                    except:
                        pass
    
    workbook.close()
    
    print(f"✅ Found {len(document_types)} document types in Excel")
    print(f"✅ Found {len(formats_data)} format definitions in Sheet2\n")
    return document_types, formats_data


def lookup_folder_id(cursor, folder_name):
    """Lookup folder ID by name"""
    if not folder_name:
        return None
    cursor.execute("SELECT id FROM folders WHERE name = %s LIMIT 1", (folder_name,))
    result = cursor.fetchone()
    return result[0] if result else None


def lookup_security_id(cursor, security_name):
    """Lookup security classification ID by name"""
    if not security_name:
        return None
    cursor.execute("SELECT id FROM security_classifications WHERE name = %s LIMIT 1", (security_name,))
    result = cursor.fetchone()
    return result[0] if result else None


def delete_existing_numbering_formats(cursor, document_type_id):
    """Delete existing numbering formats and components for a document type"""
    # Components akan terhapus otomatis karena CASCADE constraint
    cursor.execute("""
        DELETE FROM numbering_formats 
        WHERE official_document_type_id = %s
    """, (document_type_id,))


def delete_existing_document_sequences(cursor, document_type_id):
    """Delete existing document sequences for a document type"""
    cursor.execute("""
        DELETE FROM document_sequences 
        WHERE official_document_type_id = %s
    """, (document_type_id,))


def initialize_document_sequences(cursor, document_type_id, formats_data, scope_level):
    """Initialize document sequences for each format category"""
    now = datetime.now()
    current_year = now.year
    
    # Get formats based on scope_level
    formats_to_create = []
    
    if scope_level == 'CORPORATE':
        if 'CORPORATE' in formats_data:
            formats_to_create = formats_data['CORPORATE']
    elif scope_level == 'UNIT_PROJECT':
        if 'UNIT_PROJECT' in formats_data:
            formats_to_create = formats_data['UNIT_PROJECT']
    
    if not formats_to_create:
        return
    
    # Initialize sequence for each format category
    for format_data in formats_to_create:
        category = format_data.get('category')
        
        if not category:
            continue
        
        # Insert document_sequence with initial counter = 0
        sequence_id = generate_ulid()
        cursor.execute("""
            INSERT INTO document_sequences (
                id, official_document_type_id, format_category,
                year, work_unit_id, project_id,
                current_sequence, parent_sequence_number,
                created_at, updated_at
            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        """, (
            sequence_id,
            document_type_id,
            category,
            current_year,
            None,  # work_unit_id - NULL untuk master data
            None,  # project_id - NULL untuk master data
            0,     # current_sequence - mulai dari 0
            None,  # parent_sequence_number
            now,
            now
        ))


def create_numbering_formats(cursor, document_type_id, formats_data, scope_level):
    """Create numbering formats and components for a document type"""
    now = datetime.now()
    
    # Get formats based on scope_level
    formats_to_create = []
    
    if scope_level == 'CORPORATE':
        if 'CORPORATE' in formats_data:
            formats_to_create = formats_data['CORPORATE']
    elif scope_level == 'UNIT_PROJECT':
        if 'UNIT_PROJECT' in formats_data:
            formats_to_create = formats_data['UNIT_PROJECT']
    
    if not formats_to_create:
        return
    
    # Create numbering format for each category
    for format_data in formats_to_create:
        category = format_data.get('category')
        components = format_data.get('components', [])
        
        if not category:
            continue
        
        # Insert numbering_format
        format_id = generate_ulid()
        cursor.execute("""
            INSERT INTO numbering_formats (
                id, official_document_type_id, format_category, created_at, updated_at
            ) VALUES (%s, %s, %s, %s, %s)
        """, (format_id, document_type_id, category, now, now))
        
        # Insert numbering_components
        for component in components:
            component_id = generate_ulid()
            component_type = component.get('type')
            sequence = component.get('sequence')
            
            variable_key = None
            static_value = None
            
            if component_type == 'VARIABLE':
                variable_key = component.get('key')
            elif component_type == 'STATIC_TEXT':
                static_value = component.get('value')
            
            cursor.execute("""
                INSERT INTO numbering_components (
                    id, numbering_format_id, sequence_order,
                    component_type, variable_key, static_value,
                    created_at, updated_at
                ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
            """, (component_id, format_id, sequence, component_type, variable_key, static_value, now, now))


def build_template_url(filename):
    """Build full template URL from filename"""
    if not filename:
        return None
    # Remove leading slash if present
    filename = filename.lstrip('/')
    return f"{TEMPLATE_BASE_URL}/{filename}"


def upsert_document_type(cursor, data, formats_data):
    """Upsert document type (insert if not exists, update if exists)"""
    now = datetime.now()
    name = data.get('name')
    
    if not name:
        return None, False
    
    # Build full template URL from filename
    template_path = build_template_url(data.get('template_path'))
    
    # Lookup folder_id and security_classification_id
    folder_name = data.get('folder_id')  # Di Excel ini nama folder
    security_name = data.get('security_classification_id')  # Di Excel ini nama security
    
    folder_id = lookup_folder_id(cursor, folder_name) if folder_name else None
    security_id = lookup_security_id(cursor, security_name) if security_name else None
    
    # Get scope_level
    scope_level = data.get('scope_level', 'CORPORATE')
    
    # Cek apakah sudah ada
    cursor.execute("SELECT id, code FROM official_document_types WHERE name = %s", (name,))
    existing = cursor.fetchone()
    
    if existing:
        # Update existing record
        existing_id, existing_code = existing
        
        # Tentukan code value: dari Excel atau keep existing
        code_value = data.get('code') if data.get('code') else existing_code
        
        cursor.execute("""
            UPDATE official_document_types
            SET 
                updated_at = %s,
                code = %s,
                folder_id = %s,
                security_classification_id = %s,
                retention_period_years = %s,
                template_path = %s,
                extended_by_owner = %s,
                allow_multiple_documents = %s,
                has_numbering = %s,
                scope_level = %s
            WHERE id = %s
        """, (
            now,
            code_value,
            folder_id,
            security_id,
            data.get('retention_period_years', 5),
            template_path,
            data.get('extended_by_owner', False),
            data.get('allow_multiple_documents', False),
            data.get('has_numbering', True),
            scope_level,
            existing_id
        ))
        
        # Delete existing numbering formats and sequences, then recreate
        if data.get('has_numbering', True):
            delete_existing_numbering_formats(cursor, existing_id)
            delete_existing_document_sequences(cursor, existing_id)
            create_numbering_formats(cursor, existing_id, formats_data, scope_level)
            initialize_document_sequences(cursor, existing_id, formats_data, scope_level)
        
        return existing_id, True
    else:
        # Insert new record dengan ULID
        new_id = generate_ulid()
        code_value = data.get('code')
        
        cursor.execute("""
            INSERT INTO official_document_types (
                id, name, code, created_at, updated_at,
                folder_id, security_classification_id,
                letter_category_id, letter_sub_category_id,
                template_path, retention_period_years, is_active,
                extended_by_owner, allow_multiple_documents,
                has_numbering, scope_level
            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        """, (
            new_id,
            name,
            code_value,
            now,
            now,
            folder_id,
            security_id,
            None,  # letter_category_id - bisa diisi manual nanti jika diperlukan
            None,  # letter_sub_category_id
            template_path,
            data.get('retention_period_years', 5),
            True,  # is_active
            data.get('extended_by_owner', False),
            data.get('allow_multiple_documents', False),
            data.get('has_numbering', True),
            scope_level
        ))
        
        # Create numbering formats, components, and initialize sequences
        if data.get('has_numbering', True):
            create_numbering_formats(cursor, new_id, formats_data, scope_level)
            initialize_document_sequences(cursor, new_id, formats_data, scope_level)
        
        return new_id, False


def main():
    """Fungsi utama untuk inject data dari Excel"""
    print("=" * 70)
    print("Script Inject Document Types from Excel (UPSERT Mode)")
    print("=" * 70)
    
    # Baca data dari Excel
    try:
        document_types, formats_data = read_excel_data()
        if not document_types:
            print("❌ Tidak ada data untuk diproses")
            return
    except Exception as e:
        print(f"❌ Error reading Excel: {e}")
        import traceback
        traceback.print_exc()
        return
    
    # Koneksi ke database
    print("📡 Connecting to database...")
    conn = connect_db()
    if not conn:
        return
    
    cursor = conn.cursor()
    
    try:
        inserted_count = 0
        updated_count = 0
        skipped_count = 0
        
        print(f"✅ Connected!\n")
        print(f"🔄 Processing {len(document_types)} document types...\n")
        print("-" * 70)
        
        for idx, doc_type in enumerate(document_types, 1):
            name = doc_type.get('name', '')
            code = doc_type.get('code', '')
            
            if not name:
                skipped_count += 1
                continue
            
            display_info = f"{name}"
            if code:
                display_info += f" ({code})"
            
            print(f"[{idx}/{len(document_types)}] {display_info}")
            
            # Upsert
            result = upsert_document_type(cursor, doc_type, formats_data)
            
            if result[0] is None:
                print(f"   ⏭️  Skipped (no name)")
                skipped_count += 1
            else:
                doc_id, is_update = result
                
                # Count numbering formats and sequences created
                cursor.execute("""
                    SELECT COUNT(*) FROM numbering_formats 
                    WHERE official_document_type_id = %s
                """, (doc_id,))
                format_count = cursor.fetchone()[0]
                
                cursor.execute("""
                    SELECT COUNT(*) FROM document_sequences 
                    WHERE official_document_type_id = %s
                """, (doc_id,))
                sequence_count = cursor.fetchone()[0]
                
                if is_update:
                    print(f"   🔄 Updated: {doc_id}")
                    if format_count > 0:
                        print(f"      └─ Numbering formats: {format_count} format(s)")
                        print(f"      └─ Document sequences: {sequence_count} sequence(s)")
                    updated_count += 1
                else:
                    print(f"   ✅ Inserted: {doc_id}")
                    if format_count > 0:
                        print(f"      └─ Numbering formats: {format_count} format(s)")
                        print(f"      └─ Document sequences: {sequence_count} sequence(s)")
                    inserted_count += 1
        
        print("\n" + "-" * 70)
        print(f"\n📊 Summary:")
        print(f"   Total rows in Excel: {len(document_types)}")
        print(f"   New inserts:         {inserted_count}")
        print(f"   Updated:             {updated_count}")
        print(f"   Skipped:             {skipped_count}")
        
        # Count total numbering formats, components, and sequences
        cursor.execute("SELECT COUNT(*) FROM numbering_formats")
        total_formats = cursor.fetchone()[0]
        cursor.execute("SELECT COUNT(*) FROM numbering_components")
        total_components = cursor.fetchone()[0]
        cursor.execute("SELECT COUNT(*) FROM document_sequences")
        total_sequences = cursor.fetchone()[0]
        
        print(f"\n   Total numbering formats:    {total_formats}")
        print(f"   Total numbering components: {total_components}")
        print(f"   Total document sequences:   {total_sequences}")
        print("-" * 70)
        
        # Tampilkan preview
        if inserted_count > 0 or updated_count > 0:
            print("\n📋 Recent Records (top 10):")
            print("-" * 70)
            cursor.execute("""
                SELECT id, name, retention_period_years, has_numbering, scope_level, created_at
                FROM official_document_types 
                ORDER BY updated_at DESC 
                LIMIT 10
            """)
            records = cursor.fetchall()
            
            for idx, record in enumerate(records, 1):
                doc_id, doc_name, retention, has_numbering, scope_level, created_at = record
                print(f"\n   {idx}. {doc_name}")
                print(f"      ID: {doc_id}")
                print(f"      Retention: {retention} years")
                print(f"      Scope: {scope_level}")
                print(f"      Has numbering: {has_numbering}")
                
                if has_numbering:
                    # Show numbering formats for this doc type
                    cursor.execute("""
                        SELECT nf.format_category, COUNT(nc.id) as component_count
                        FROM numbering_formats nf
                        LEFT JOIN numbering_components nc ON nc.numbering_format_id = nf.id
                        WHERE nf.official_document_type_id = %s
                        GROUP BY nf.format_category
                        ORDER BY nf.format_category
                    """, (doc_id,))
                    formats = cursor.fetchall()
                    if formats:
                        print(f"      Formats: {', '.join([f'{cat}({cnt})' for cat, cnt in formats])}")
                
                print(f"      Created: {created_at}")
            
            print("\n" + "-" * 70)
        
        # Konfirmasi commit
        print("\n⚠️  Review data di atas sebelum commit!")
        commit_choice = input("Commit ke database? (yes/no): ").lower()
        
        if commit_choice in ['yes', 'y']:
            conn.commit()
            print("\n" + "=" * 70)
            print("✅ SUCCESS - Data berhasil disimpan!")
            print("=" * 70)
        else:
            conn.rollback()
            print("\n" + "=" * 70)
            print("❌ ROLLBACK - Perubahan dibatalkan")
            print("=" * 70)
    
    except Exception as e:
        conn.rollback()
        print(f"\n❌ Error: {e}")
        print("Transaksi di-rollback")
        import traceback
        traceback.print_exc()
    
    finally:
        cursor.close()
        conn.close()
        print("\n✅ Koneksi database ditutup\n")


if __name__ == "__main__":
    # Cek package
    try:
        import psycopg2
        from ulid import ULID
        from openpyxl import load_workbook
    except ImportError as e:
        print("❌ Package belum terinstall!")
        print("\nInstall dengan:")
        print("pip install psycopg2-binary ulid-py openpyxl")
        exit(1)
    
    print(f"\n📂 Excel File: {EXCEL_FILE}")
    print(f"🔗 Database: {DATABASE_URL}\n")
    
    proceed = input("Lanjutkan inject dari Excel? (yes/no): ").lower()
    if proceed in ['yes', 'y']:
        main()
    else:
        print("❌ Dibatalkan")
